"""
Report generation service for PDF and Excel exports.
"""
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta
from sqlalchemy.orm import Session
from sqlalchemy import func, extract, desc, distinct
from io import BytesIO
from collections import Counter
import re
import logging

logger = logging.getLogger(__name__)

from models.patient import Patient
from models.medecin import Medecin
from models.rendez_vous import RendezVous
from models.consultation import Consultation


def generer_rapport_global(db: Session) -> dict:
    """
    Generate global statistics report.
    """
    total_patients = db.query(func.count(Patient.id)).scalar()
    total_medecins = db.query(func.count(Medecin.id)).scalar()
    total_rendez_vous = db.query(func.count(RendezVous.id)).scalar()
    
    # Patients this month
    current_month = datetime.now().month
    current_year = datetime.now().year
    nouveaux_patients = db.query(func.count(Patient.id)).filter(
        extract('month', Patient.created_at) == current_month,
        extract('year', Patient.created_at) == current_year
    ).scalar()

    # Rendez-vous today
    today = datetime.now().date()
    rdv_today = db.query(func.count(RendezVous.id)).filter(
        func.date(RendezVous.date_heure) == today
    ).scalar()
    
    # Count appointments by status
    rendez_vous_par_statut = db.query(
        RendezVous.statut,
        func.count(RendezVous.id)
    ).group_by(RendezVous.statut).all()
    
    # Most requested specialties
    specialites = db.query(
        Medecin.specialite,
        func.count(RendezVous.id).label('count')
    ).join(RendezVous, Medecin.id == RendezVous.medecin_id)\
     .group_by(Medecin.specialite)\
     .order_by(func.count(RendezVous.id).desc())\
     .limit(10)\
     .all()
    
    return {
        "total_patients": total_patients,
        "total_medecins": total_medecins,
        "total_rendez_vous": total_rendez_vous,
        "nouveaux_patients_mois": nouveaux_patients,
        "rdv_aujourdhui": rdv_today,
        "rendez_vous_par_statut": [
            {"statut": statut.value if hasattr(statut, 'value') else str(statut), "count": count} 
            for statut, count in rendez_vous_par_statut
        ],
        "specialites_populaires": [
            {"specialite": spec, "count": count} 
            for spec, count in specialites
        ],
        "date_generation": datetime.now().isoformat()
    }


def generer_stats_medecins(db: Session) -> list:
    """Returns statistics for each doctor."""
    try:
        medecins = db.query(Medecin).all()
        stats = []
        
        # Fixed consultation price
        CONSULTATION_PRICE = 50
        
        for med in medecins:
            # Consultations count from Consultation table (actual consultations)
            nb_consultations = db.query(func.count(Consultation.id)).filter(
                Consultation.medecin_id == med.id
            ).scalar()
            
            # Patients count (Distinct patients seen in consultations)
            nb_patients = db.query(func.count(distinct(Consultation.dossier_id))).filter(
                Consultation.medecin_id == med.id
            ).scalar()
            
            # Time-based consultation counts
            current_month = datetime.now().month
            current_year = datetime.now().year
            today = datetime.now().date()
            start_of_week = today - timedelta(days=today.weekday())
            
            # Consultations this month
            nb_consultations_mois = db.query(func.count(Consultation.id)).filter(
                Consultation.medecin_id == med.id,
                extract('month', Consultation.date_creation) == current_month,
                extract('year', Consultation.date_creation) == current_year
            ).scalar()
            
            # Consultations this week
            nb_consultations_semaine = db.query(func.count(Consultation.id)).filter(
                Consultation.medecin_id == med.id,
                func.date(Consultation.date_creation) >= start_of_week
            ).scalar()
            
            # Consultations today
            nb_consultations_aujourdhui = db.query(func.count(Consultation.id)).filter(
                Consultation.medecin_id == med.id,
                func.date(Consultation.date_creation) == today
            ).scalar()
            
            # Calculate revenues
            revenus_total = nb_consultations * CONSULTATION_PRICE
            revenus_mois = nb_consultations_mois * CONSULTATION_PRICE
            revenus_semaine = nb_consultations_semaine * CONSULTATION_PRICE
            revenus_aujourdhui = nb_consultations_aujourdhui * CONSULTATION_PRICE
            
            stats.append({
                "id": med.id,
                "nom": med.nom,
                "prenom": med.prenom,
                "specialite": med.specialite,
                "nb_consultations": nb_consultations,
                "nb_consultations_mois": nb_consultations_mois,
                "nb_consultations_semaine": nb_consultations_semaine,
                "nb_consultations_aujourdhui": nb_consultations_aujourdhui,
                "nb_patients": nb_patients,
                "revenus_total": revenus_total,
                "revenus_mois": revenus_mois,
                "revenus_semaine": revenus_semaine,
                "revenus_aujourdhui": revenus_aujourdhui
            })
        
        # Sort by consultations desc
        stats.sort(key=lambda x: x['nb_consultations'], reverse=True)
        return stats
    except Exception as e:
        logger.error(f"Error in generer_stats_medecins: {e}")
        return []


def generer_stats_rendez_vous(db: Session) -> dict:
    """Returns advanced aptitude statistics."""
    try:
        # 1. Busy Days (MySQL specific: DAYOFWEEK 1=Sunday, 7=Saturday)
        busy_days_query = db.query(
            func.dayofweek(RendezVous.date_heure).label('dow'),
            func.count(RendezVous.id)
        ).group_by('dow').all()
        
        days_map = {1: 'Dimanche', 2: 'Lundi', 3: 'Mardi', 4: 'Mercredi', 5: 'Jeudi', 6: 'Vendredi', 7: 'Samedi'}
        
        jours_charges = []
        for dow, count in busy_days_query:
            try:
                if dow is None: continue
                day_idx = int(dow)
                formatted_day = days_map.get(day_idx, str(day_idx))
                jours_charges.append({"jour": formatted_day, "count": count})
            except:
                continue
                
        # 2. Peak Hours
        busy_hours_query = db.query(
            func.hour(RendezVous.date_heure).label('hour'),
            func.count(RendezVous.id)
        ).group_by('hour').order_by('hour').all()
        
        heures_pointe = [{"heure": f"{int(h)}h", "count": c} for h, c in busy_hours_query if h is not None]
        
        # 3. Cancellation Rate
        total = db.query(func.count(RendezVous.id)).scalar()
        cancelled = db.query(func.count(RendezVous.id)).filter(RendezVous.statut == 'Annulé').scalar()
        
        taux_annulation = (cancelled / total * 100) if total and total > 0 else 0
        
        return {
            "jours_charges": jours_charges,
            "heures_pointe": heures_pointe,
            "taux_annulation": round(taux_annulation, 2)
        }
    except Exception as e:
        logger.error(f"Error in generer_stats_rendez_vous: {e}")
        return {
            "jours_charges": [],
            "heures_pointe": [],
            "taux_annulation": 0
        }


def generer_stats_medicales(db: Session) -> dict:
    """Returns frequency analysis of medical terms."""
    try:
        # 1. Top Pathologies (based on Consultation.titre)
        pathologies = db.query(
            Consultation.titre,
            func.count(Consultation.id)
        ).filter(Consultation.titre != None)\
         .group_by(Consultation.titre)\
         .order_by(func.count(Consultation.id).desc())\
         .limit(10).all()
         
        # 2. Treatments and Allergies (Token analysis)
        consultations = db.query(Consultation.traitements, Consultation.allergies).all()
        
        traitement_words = []
        allergie_words = []
        
        for trait, allerg in consultations:
            if trait:
                parts = re.split(r'[,\n\r]+', trait)
                traitement_words.extend([p.strip().lower() for p in parts if p.strip()])
                
            if allerg:
                parts = re.split(r'[,\n\r]+', allerg)
                allergie_words.extend([p.strip().lower() for p in parts if p.strip()])
                
        top_traitements = Counter(traitement_words).most_common(10)
        top_allergies = Counter(allergie_words).most_common(10)
        
        return {
            "pathologies_frequentes": [{"nom": t, "count": c} for t, c in pathologies],
            "traitements_populaires": [{"nom": t.title(), "count": c} for t, c in top_traitements],
            "allergies_frequentes": [{"nom": t.title(), "count": c} for t, c in top_allergies]
        }
    except Exception as e:
        logger.error(f"Error in generer_stats_medicales: {e}")
        return {
            "pathologies_frequentes": [],
            "traitements_populaires": [],
            "allergies_frequentes": []
        }


def generer_pdf_rapport(db: Session) -> BytesIO:
    """
    Generate PDF report.
    
    Args:
        db: Database session
        
    Returns:
        BytesIO buffer containing the PDF
    """
    try:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=1  # Center
        )
        elements.append(Paragraph("Rapport Global - Clinique Médicale", title_style))
        elements.append(Spacer(1, 0.3 * inch))
        
        # Get all statistics
        stats = generer_rapport_global(db)
        med_stats = generer_stats_medecins(db)
        rdv_stats = generer_stats_rendez_vous(db)
        medical_stats = generer_stats_medicales(db)
        
        # Date
        date_text = f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        elements.append(Paragraph(date_text, styles['Normal']))
        elements.append(Spacer(1, 0.3 * inch))
        
        # 1. Global Statistics Table
        elements.append(Paragraph("1. Statistiques Générales", styles['Heading2']))
        elements.append(Spacer(1, 0.1 * inch))
        
        data = [
            ['Statistique', 'Valeur'],
            ['Total Patients', str(stats['total_patients'])],
            ['Total Médecins', str(stats['total_medecins'])],
            ['Total Rendez-vous', str(stats['total_rendez_vous'])],
            ['Patients ce mois', str(stats['nouveaux_patients_mois'])],
            ['RDV Aujourd\'hui', str(stats['rdv_aujourdhui'])]
        ]
        
        table = Table(data, colWidths=[3 * inch, 2 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4A90E2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.4 * inch))
        
        # 1.5 Specialties (Added based on user request)
        if stats['specialites_populaires']:
            elements.append(Paragraph("Spécialités les plus sollicitées", styles['Heading2']))
            elements.append(Spacer(1, 0.1 * inch))
            
            spec_data = [['Spécialité', 'Rendez-vous']]
            for spec in stats['specialites_populaires']:
                spec_data.append([spec['specialite'], str(spec['count'])])
                
            spec_table = Table(spec_data, colWidths=[3 * inch, 2 * inch])
            spec_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9C27B0')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(spec_table)
            elements.append(Spacer(1, 0.4 * inch))
        
        # 2. Doctor Performance
        if med_stats:
            elements.append(Paragraph("2. Performance Médecins", styles['Heading2']))
            elements.append(Spacer(1, 0.1 * inch))
            
            med_data = [['Médecin', 'Spécialité', 'Patients', 'Consultations', 'Ce Mois']]
            for med in med_stats:
                med_data.append([
                    f"Dr. {med['nom']}", 
                    med['specialite'], 
                    str(med['nb_patients']), 
                    str(med['nb_consultations']),
                    str(med.get('nb_consultations_mois', 0))
                ])
                
            med_table = Table(med_data, colWidths=[2 * inch, 1.5 * inch, 0.8 * inch, 1.2 * inch, 0.8 * inch])
            med_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#50C878')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(med_table)
            elements.append(Spacer(1, 0.4 * inch))
            
        # 3. Appointment Analytics
        elements.append(Paragraph("3. Analyse Rendez-vous", styles['Heading2']))
        elements.append(Spacer(1, 0.1 * inch))
        
        elements.append(Paragraph(f"<b>Taux d'annulation:</b> {rdv_stats['taux_annulation']}%", styles['Normal']))
        
        if rdv_stats['jours_charges']:
            busiest_day = rdv_stats['jours_charges'][0]
            elements.append(Paragraph(f"<b>Jour le plus chargé:</b> {busiest_day['jour']} ({busiest_day['count']} RDVs)", styles['Normal']))
            
        if rdv_stats['heures_pointe']:
            # Safe peak finding
            try:
                peak = max(rdv_stats['heures_pointe'], key=lambda x: x['count']) if rdv_stats['heures_pointe'] else None
                if peak:
                    elements.append(Paragraph(f"<b>Heure de pointe:</b> {peak['heure']} ({peak['count']} RDVs)", styles['Normal']))
            except Exception as e:
                logger.error(f"Error calculating peak hour: {e}")
                
        elements.append(Spacer(1, 0.4 * inch))

        # 4. Medical Insights
        elements.append(Paragraph("4. Statistiques Médicales", styles['Heading2']))
        elements.append(Spacer(1, 0.1 * inch))
        
        if medical_stats['pathologies_frequentes']:
            elements.append(Paragraph("<b>Pathologies Fréquentes:</b>", styles['Normal']))
            path_text = ", ".join([f"{p['nom']} ({p['count']})" for p in medical_stats['pathologies_frequentes'][:5]])
            elements.append(Paragraph(path_text, styles['Normal']))
            elements.append(Spacer(1, 0.1 * inch))

        if medical_stats['traitements_populaires']:
            elements.append(Paragraph("<b>Traitements Top 5:</b>", styles['Normal']))
            trait_text = ", ".join([f"{t['nom']} ({t['count']})" for t in medical_stats['traitements_populaires'][:5]])
            elements.append(Paragraph(trait_text, styles['Normal']))

        doc.build(elements)
        buffer.seek(0)
        return buffer
    except Exception as e:
        logger.error(f"Error generating PDF: {e}")
        # Return empty buffer or re-raise depending on need, 
        # but re-raising lets the API return 500 which is caught by user.
        # Let's print to console for debug
        print(f"CRITICAL ERROR generating PDF: {e}")
        raise e


def generer_excel_rapport(db: Session) -> BytesIO:
    """
    Generate Excel report.
    
    Args:
        db: Database session
        
    Returns:
        BytesIO buffer containing the Excel file
    """
    buffer = BytesIO()
    wb = Workbook()
    
    # Statistics sheet
    ws_stats = wb.active
    ws_stats.title = "Statistiques Globales"
    
    # Header styling
    header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Get statistics
    stats = generer_rapport_global(db)
    
    # Write headers
    ws_stats['A1'] = "Rapport Clinique Médicale"
    ws_stats['A1'].font = Font(bold=True, size=16)
    ws_stats['A2'] = f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    
    # Statistics
    ws_stats['A4'] = "Statistique"
    ws_stats['B4'] = "Valeur"
    ws_stats['A4'].fill = header_fill
    ws_stats['B4'].fill = header_fill
    ws_stats['A4'].font = header_font
    ws_stats['B4'].font = header_font
    
    ws_stats['A5'] = "Total Patients"
    ws_stats['B5'] = stats['total_patients']
    ws_stats['A6'] = "Total Médecins"
    ws_stats['B6'] = stats['total_medecins']
    ws_stats['A7'] = "Total Rendez-vous"
    ws_stats['B7'] = stats['total_rendez_vous']
    ws_stats['A8'] = "Patients ce mois"
    ws_stats['B8'] = stats['nouveaux_patients_mois']
    ws_stats['A9'] = "RDV Aujourd'hui"
    ws_stats['B9'] = stats['rdv_aujourdhui']
    
    # Specialties sheet
    if stats['specialites_populaires']:
        ws_spec = wb.create_sheet("Spécialités")
        ws_spec['A1'] = "Spécialité"
        ws_spec['B1'] = "Nombre de rendez-vous"
        ws_spec['A1'].fill = header_fill
        ws_spec['B1'].fill = header_fill
        ws_spec['A1'].font = header_font
        ws_spec['B1'].font = header_font
        
        for idx, spec in enumerate(stats['specialites_populaires'], start=2):
            ws_spec[f'A{idx}'] = spec['specialite']
            ws_spec[f'B{idx}'] = spec['count']
    
    # Adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generer_rapport_patients_pdf(db: Session) -> BytesIO:
    """
    Generate PDF report of all patients.
    """
    try:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=1
        )
        elements.append(Paragraph("Liste des Patients", title_style))
        elements.append(Spacer(1, 0.3 * inch))
        
        # Date
        date_text = f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        elements.append(Paragraph(date_text, styles['Normal']))
        elements.append(Spacer(1, 0.3 * inch))
        
        # Table Data
        patients = db.query(Patient).order_by(Patient.nom, Patient.prenom).all()
        
        data = [['Nom', 'Prénom', 'Téléphone', 'Email', 'N° SS']]
        for p in patients:
            data.append([
                p.nom,
                p.prenom,
                p.telephone or "-",
                p.email or "-",
                p.numero_securite_sociale or "-"
            ])
            
        table = Table(data, colWidths=[1.5*inch, 1.5*inch, 1.2*inch, 2*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4A90E2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.beige])
        ]))
        
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        return buffer
    except Exception as e:
        logger.error(f"Error generating Patients PDF: {e}")
        raise e


def generer_rapport_patients_excel(db: Session) -> BytesIO:
    """
    Generate Excel report of all patients.
    """
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Liste Patients"
    
    # Headers
    headers = ['ID', 'Nom', 'Prénom', 'Date Naissance', 'Sexe', 'Téléphone', 'Email', 'Adresse', 'N° SS', 'Date Création']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
    
    # Data
    patients = db.query(Patient).order_by(Patient.nom, Patient.prenom).all()
    for row_idx, p in enumerate(patients, 2):
        ws.cell(row=row_idx, column=1, value=p.id)
        ws.cell(row=row_idx, column=2, value=p.nom)
        ws.cell(row=row_idx, column=3, value=p.prenom)
        ws.cell(row=row_idx, column=4, value=str(p.date_naissance) if p.date_naissance else "-")
        ws.cell(row=row_idx, column=5, value=p.sexe)
        ws.cell(row=row_idx, column=6, value=p.telephone or "-")
        ws.cell(row=row_idx, column=7, value=p.email or "-")
        ws.cell(row=row_idx, column=8, value=p.adresse or "-")
        ws.cell(row=row_idx, column=9, value=p.numero_securite_sociale or "-")
        ws.cell(row=row_idx, column=10, value=str(p.created_at))
        
    # Auto-width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
        
    wb.save(buffer)
    buffer.seek(0)
    return buffer
