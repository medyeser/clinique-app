"""
Revenue report generation service for PDF and Excel exports.
"""
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, extract
from io import BytesIO
import logging

logger = logging.getLogger(__name__)

from models.consultation import Consultation
from models.medecin import Medecin
from models.patient import Patient
from models.dossier_medical import DossierMedical

CONSULTATION_PRICE = 50


def generer_rapport_revenus_pdf(db: Session, start_date_str: str, end_date_str: str) -> BytesIO:
    """
    Generate PDF revenue report for a date range.
    """
    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=1  # Center
        )
        elements.append(Paragraph("Rapport de Revenus", title_style))
        elements.append(Spacer(1, 0.3 * inch))
        
        # Date range
        date_text = f"Période: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
        elements.append(Paragraph(date_text, styles['Normal']))
        gen_text = f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        elements.append(Paragraph(gen_text, styles['Normal']))
        elements.append(Spacer(1, 0.3 * inch))
        
        # Get consultations
        consultations = db.query(Consultation).filter(
            and_(
                func.date(Consultation.date_creation) >= start_date,
                func.date(Consultation.date_creation) <= end_date
            )
        ).order_by(Consultation.date_creation).all()
        
        total_consultations = len(consultations)
        total_revenue = total_consultations * CONSULTATION_PRICE
        
        # Summary section
        elements.append(Paragraph("Résumé", styles['Heading2']))
        elements.append(Spacer(1, 0.1 * inch))
        
        summary_data = [
            ['Indicateur', 'Valeur'],
            ['Nombre total de consultations', str(total_consultations)],
            ['Revenu total', f'{total_revenue} DT'],
            ['Prix par consultation', f'{CONSULTATION_PRICE} DT'],
            ['Revenu moyen par jour', f'{total_revenue / max((end_date - start_date).days + 1, 1):.2f} DT']
        ]
        
        summary_table = Table(summary_data, colWidths=[3.5 * inch, 2 * inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4A90E2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 0.4 * inch))
        
        # Revenue by doctor
        elements.append(Paragraph("Revenus par Médecin", styles['Heading2']))
        elements.append(Spacer(1, 0.1 * inch))
        
        # Calculate revenue by doctor
        medecin_revenues = {}
        for cons in consultations:
            if cons.medecin_id:
                if cons.medecin_id not in medecin_revenues:
                    medecin = db.query(Medecin).filter(Medecin.id == cons.medecin_id).first()
                    medecin_revenues[cons.medecin_id] = {
                        "nom": f"Dr. {medecin.nom} {medecin.prenom}" if medecin else "Inconnu",
                        "specialite": medecin.specialite if medecin else "",
                        "consultations": 0,
                        "revenue": 0
                    }
                medecin_revenues[cons.medecin_id]["consultations"] += 1
                medecin_revenues[cons.medecin_id]["revenue"] += CONSULTATION_PRICE
        
        if medecin_revenues:
            med_data = [['Médecin', 'Spécialité', 'Consultations', 'Revenus (DT)']]
            for med_info in sorted(medecin_revenues.values(), key=lambda x: x['revenue'], reverse=True):
                med_data.append([
                    med_info['nom'],
                    med_info['specialite'],
                    str(med_info['consultations']),
                    f"{med_info['revenue']} DT"
                ])
            
            med_table = Table(med_data, colWidths=[2 * inch, 1.5 * inch, 1.2 * inch, 1.3 * inch])
            med_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#50C878')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.beige])
            ]))
            elements.append(med_table)
        else:
            elements.append(Paragraph("Aucune consultation enregistrée pour cette période.", styles['Normal']))
        
        doc.build(elements)
        buffer.seek(0)
        return buffer
    except Exception as e:
        logger.error(f"Error generating revenue PDF: {e}")
        raise e


def generer_rapport_revenus_excel(db: Session, start_date_str: str, end_date_str: str) -> BytesIO:
    """
    Generate Excel revenue report for a date range.
    """
    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        
        buffer = BytesIO()
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Résumé"
        
        # Header styling
        header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Title
        ws_summary['A1'] = "Rapport de Revenus"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary['A2'] = f"Période: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
        ws_summary['A3'] = f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        
        # Get consultations
        consultations = db.query(Consultation).filter(
            and_(
                func.date(Consultation.date_creation) >= start_date,
                func.date(Consultation.date_creation) <= end_date
            )
        ).order_by(Consultation.date_creation).all()
        
        total_consultations = len(consultations)
        total_revenue = total_consultations * CONSULTATION_PRICE
        
        # Summary data
        ws_summary['A5'] = "Indicateur"
        ws_summary['B5'] = "Valeur"
        ws_summary['A5'].fill = header_fill
        ws_summary['B5'].fill = header_fill
        ws_summary['A5'].font = header_font
        ws_summary['B5'].font = header_font
        
        ws_summary['A6'] = "Nombre total de consultations"
        ws_summary['B6'] = total_consultations
        ws_summary['A7'] = "Revenu total (DT)"
        ws_summary['B7'] = total_revenue
        ws_summary['A8'] = "Prix par consultation (DT)"
        ws_summary['B8'] = CONSULTATION_PRICE
        
        # Revenue by doctor sheet
        ws_medecins = wb.create_sheet("Revenus par Médecin")
        ws_medecins['A1'] = "Médecin"
        ws_medecins['B1'] = "Spécialité"
        ws_medecins['C1'] = "Consultations"
        ws_medecins['D1'] = "Revenus (DT)"
        
        for col in ['A1', 'B1', 'C1', 'D1']:
            ws_medecins[col].fill = header_fill
            ws_medecins[col].font = header_font
        
        # Calculate revenue by doctor
        medecin_revenues = {}
        for cons in consultations:
            if cons.medecin_id:
                if cons.medecin_id not in medecin_revenues:
                    medecin = db.query(Medecin).filter(Medecin.id == cons.medecin_id).first()
                    medecin_revenues[cons.medecin_id] = {
                        "nom": f"Dr. {medecin.nom} {medecin.prenom}" if medecin else "Inconnu",
                        "specialite": medecin.specialite if medecin else "",
                        "consultations": 0,
                        "revenue": 0
                    }
                medecin_revenues[cons.medecin_id]["consultations"] += 1
                medecin_revenues[cons.medecin_id]["revenue"] += CONSULTATION_PRICE
        
        row = 2
        for med_info in sorted(medecin_revenues.values(), key=lambda x: x['revenue'], reverse=True):
            ws_medecins[f'A{row}'] = med_info['nom']
            ws_medecins[f'B{row}'] = med_info['specialite']
            ws_medecins[f'C{row}'] = med_info['consultations']
            ws_medecins[f'D{row}'] = med_info['revenue']
            row += 1
        
        # Detailed consultations sheet
        ws_details = wb.create_sheet("Détails Consultations")
        headers = ['Date', 'Heure', 'Médecin', 'Spécialité', 'Patient', 'Titre', 'Revenu (DT)']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_details.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        row = 2
        for cons in consultations:
            medecin = db.query(Medecin).filter(Medecin.id == cons.medecin_id).first() if cons.medecin_id else None
            dossier = db.query(DossierMedical).filter(DossierMedical.id == cons.dossier_id).first()
            patient = db.query(Patient).filter(Patient.id == dossier.patient_id).first() if dossier else None
            
            ws_details.cell(row=row, column=1, value=cons.date_creation.strftime('%d/%m/%Y'))
            ws_details.cell(row=row, column=2, value=cons.date_creation.strftime('%H:%M'))
            ws_details.cell(row=row, column=3, value=f"Dr. {medecin.nom} {medecin.prenom}" if medecin else "Non assigné")
            ws_details.cell(row=row, column=4, value=medecin.specialite if medecin else "")
            ws_details.cell(row=row, column=5, value=f"{patient.nom} {patient.prenom}" if patient else "Inconnu")
            ws_details.cell(row=row, column=6, value=cons.titre or "Consultation")
            ws_details.cell(row=row, column=7, value=CONSULTATION_PRICE)
            row += 1
        
        # Auto-width for all sheets
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    except Exception as e:
        logger.error(f"Error generating revenue Excel: {e}")
        raise e
